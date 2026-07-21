import stanza
import re
import pandas as pd
from pathlib import Path
from openpyxl.styles import Alignment, Font

# Inisialisasi fungsi alignment di awal agar tidak error
def wrap_alignment():
    return Alignment(wrap_text=True, vertical='top')

print("Memuat Pipeline Stanza...")
nlp = stanza.Pipeline('id', processors='tokenize,pos,lemma,depparse', verbose=False, use_gpu=True)

ACTIVE_MARKERS = {"mengkritik", "menyindir", "menolak", "mengecam", "menegaskan", "menyatakan", "mengatakan", "menuding", "menyerang", "membela", "menilai", "mengaku", "mengklaim", "mengimbau", "mengingatkan", "menyampaikan", "menjelaskan", "menambahkan"}
PASSIVE_MARKERS = {"dikecam", "dikritik", "dipuji", "ditahan", "dipecat", "dituding", "dituduh", "dilaporkan", "dicekal", "disindir"}
PRONOUNS = {"dia", "ia", "beliau", "mereka", "nya"}

def get_new_smart_context(text, target_entity):
    doc = nlp(text)
    sentences = []
    for sent in doc.sentences:
        if len(sent.text.strip()) > 10:
            sentences.append({
                "text": sent.text,
                "start": sent.tokens[0].start_char,
                "end": sent.tokens[-1].end_char,
                "parsed": sent
            })
            
    # Cari offset nama tokoh (atau aliasnya)
    match = re.search(re.escape(target_entity), text, re.IGNORECASE)
    if not match:
        first_name = target_entity.split()[0]
        match = re.search(r'\b' + re.escape(first_name) + r'\b', text, re.IGNORECASE)
        if not match:
            return "Tidak ditemukan", []
        
    start_offset = match.start()
    
    anchor_idx = -1
    for i, s in enumerate(sentences):
        if s["start"] <= start_offset < s["end"]:
            anchor_idx = i
            break
            
    if anchor_idx == -1:
        return "Offset gagal dipetakan", []
        
    anchor_sent = sentences[anchor_idx]
    context_parts = []
    
    root_word = ""
    has_action = False
    for word in anchor_sent["parsed"].words:
        if word.deprel == 'root':
            root_word = (word.lemma or word.text).lower()
            if root_word in ACTIVE_MARKERS or root_word in PASSIVE_MARKERS:
                has_action = True

    # Pronoun Hunt
    if has_action and anchor_idx + 1 < len(sentences):
        next_sent = sentences[anchor_idx + 1]
        first_word = next_sent["parsed"].words[0].text.lower()
        if first_word in PRONOUNS:
            context_parts.append("[PRONOUN HUNT] " + next_sent["text"])
                
    if anchor_idx > 0:
        context_parts.append(sentences[anchor_idx - 1]["text"])
    
    context_parts.append(anchor_sent["text"])
    
    return " ".join(context_parts), context_parts

# --- BACA DATASET ---
excel_path = Path(__file__).resolve().parents[2] / "devtools" / "eval" / "finetune_dataset_gold.xlsx"
output_path = Path(__file__).resolve().parents[2] / "devtools" / "eval" / "benchmark_contexts.xlsx"

print(f"Membaca dataset dari: {excel_path}")
df = pd.read_excel(excel_path)

# Ambil 20 baris pertama untuk diuji
sample_df = df.head(20).copy()

results_data = []

for idx, row in sample_df.iterrows():
    article_text = str(row.get("article_text") or "")
    entity_name = str(row.get("entity_name") or "")
    old_ctx = str(row.get("context_text") or "")
    
    if not article_text or len(article_text) < 50:
        continue
        
    new_ctx, parts = get_new_smart_context(article_text, entity_name)
    
    results_data.append({
        "Tokoh": entity_name,
        "Artikel Asli (Head)": article_text[:100] + "...",
        "Konteks Lama (v8)": old_ctx,
        "Konteks Baru Stanza (v12)": new_ctx
    })

# --- EXPORT KE EXCEL ---
df_out = pd.DataFrame(results_data)

try:
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_out.to_excel(writer, index=False, sheet_name='Benchmark')
        
        workbook = writer.book
        worksheet = writer.sheets['Benchmark']
        
        # Buat Header Bold
        for col in worksheet.iter_cols(1, worksheet.max_column, 1, 1):
            for cell in col:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Atur lebar kolom
        col_widths = {'A': 25, 'B': 50, 'C': 80, 'D': 80}
        for col_letter, width in col_widths.items():
            worksheet.column_dimensions[col_letter].width = width
            
        # Wrap text untuk kolom konteks (Kolom C dan D)
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=3, max_col=4):
            for cell in row:
                cell.alignment = wrap_alignment()
                
    print(f"\nSELESAI! File benchmark telah disimpan di: {output_path}")
    print("Silakan buka file tersebut untuk melihat perbandingan secara utuh.")
    
except Exception as e:
    print(f"Error saat menulis Excel: {e}")