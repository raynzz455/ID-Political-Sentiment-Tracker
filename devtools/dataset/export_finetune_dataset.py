"""
export_finetune_dataset.py v8 — Auto Versioning, Excel & AI-Optimized JSONL
=====================================================================
FIX v8:
  1. AUTO VERSIONING: Mendeteksi versi terakhir di folder dan auto-increment (V4, V5, dst).
  2. SOURCE URL ADDED: Menyertakan link berita asli untuk kebutuhan audit/tracing.
  3. DUAL EXPORT:
     - Excel (.xlsx) untuk Human Labeler (Auto-fit, Wrap text, URL bisa diklik).
     - JSONL (.jsonl) untuk AI Fine-Tuning (Ringkas, hemat token, article_text dipotong 1000 chars).
"""

import sys
import math
import re
import json
import logging
import argparse
import random
from collections import Counter
from pathlib import Path
from dotenv import load_dotenv

ROOT_DIR = Path(__file__).resolve().parents[2]
load_dotenv(ROOT_DIR / ".env")
sys.path.append(str(ROOT_DIR))

logging.basicConfig(level=logging.INFO, format="[%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

from packages.shared.db_client import get_client

try:
    import pandas as pd
    from openpyxl.styles import Font, Alignment
except ImportError:
    logger.error("Dependency missing: pip install pandas openpyxl")
    sys.exit(1)

# Direktori output dataset
DATASET_DIR = ROOT_DIR / "devtools" / "dataset"
DATASET_DIR.mkdir(parents=True, exist_ok=True)

# Konfigurasi
MAX_PER_MEDIA = 500
MAX_PER_ENTITY = 300
MIN_ARTICLE_LEN = 300
MIN_CONTEXT_LEN = 50
MAX_BOILERPLATE_RATIO = 0.20
MAX_ARTICLE_CHARS_AI = 1000 # Batas karakter article_text untuk JSONL (hemat token AI)

BOILERPLATE_RE = re.compile(r'(Baca Juga|Simak Juga|Berita Terkait|Advertisement|Ikuti Kami|Copyright|©|Reportase:|Jurnalis:|Editor:).*?(?=\n|$)', re.IGNORECASE)

def get_next_version_dir(base_dir: Path) -> Path:
    """Cek folder V1, V2... lalu buat folder V(n+1) berikutnya."""
    version = 1
    while True:
        v_dir = base_dir / f"V{version}"
        if not v_dir.exists():
            v_dir.mkdir(parents=True, exist_ok=True)
            return v_dir
        version += 1

def calculate_entropy(scores: list) -> float:
    try:
        return -sum(p * math.log(p) for p in scores if p > 1e-9)
    except:
        return 1.0

def main(limit: int = 10000):
    sb = get_client()
    
    # Buat folder versi baru (misal: V4)
    version_dir = get_next_version_dir(DATASET_DIR)
    output_xlsx = version_dir / f"finetune_dataset_human.xlsx"
    output_jsonl = version_dir / f"finetune_dataset_ai.jsonl"
    
    logger.info(f"MEMULAI GOLD STANDARD CURATION (v8 Dual Export) ke folder: {version_dir.name}...")
    
    audit = {
        "raw_contexts": 0, "missing_sentiment": 0, "article_short": 0,
        "boilerplate_fail": 0, "context_short": 0, "duplicate_pair": 0,
        "entity_limit": 0, "media_limit": 0
    }
    
    try:
        ctx_res = sb.table("entity_contexts") \
                .select(
                    "raw_text_id, entity_id, context_text, metadata, "
                    "raw_texts(source_url, resolved_domain, published_at, content_hash, text), "
                    "political_entities(canonical_name)"
                ) \
                .not_.is_("entity_id", "null") \
                .limit(limit) \
                .execute()
        raw_data = ctx_res.data or []
        audit["raw_contexts"] = len(raw_data)
        logger.info(f"Total kandidat context ditarik: {len(raw_data)}")
        
        if not raw_data: return
        
        rt_ids = list(set([r["raw_text_id"] for r in raw_data]))
        ss_data = []
        for i in range(0, len(rt_ids), 100):
            chunk = rt_ids[i:i+100]
            ss_res = sb.table("sentiment_scores") \
                        .select("raw_text_id, entity_id, label, confidence, score_negative, score_neutral, score_positive") \
                        .in_("raw_text_id", chunk) \
                        .not_.is_("entity_id", "null") \
                        .execute()
            ss_data.extend(ss_res.data or [])
            
        ss_map = {(s["raw_text_id"], s["entity_id"]): s for s in ss_data}
        
    except Exception as e:
        logger.error(f"Gagal query DB: {e}")
        return

    qualified_data = []
    
    for row in raw_data:
        rt = row.get("raw_texts")
        pe = row.get("political_entities")
        if not rt or not pe: continue
        
        ss = ss_map.get((row["raw_text_id"], row["entity_id"]))
        if not ss: 
            audit["missing_sentiment"] += 1
            continue
        
        full_text = rt.get("text") or ""
        ctx_text = row.get("context_text") or ""
        meta = row.get("metadata") or {}
        entity_name = pe.get("canonical_name") or "Unknown"
        source_url = rt.get("source_url") or ""
        
        if len(full_text) < MIN_ARTICLE_LEN: 
            audit["article_short"] += 1
            continue
        
        boilerplate_hits = len(BOILERPLATE_RE.findall(full_text))
        boilerplate_ratio = (boilerplate_hits * 50) / len(full_text)
        if boilerplate_ratio > MAX_BOILERPLATE_RATIO: 
            audit["boilerplate_fail"] += 1
            continue
        
        if len(ctx_text) < MIN_CONTEXT_LEN: 
            audit["context_short"] += 1
            continue
            
        scores = [ss.get("score_negative", 0), ss.get("score_neutral", 0), ss.get("score_positive", 0)]
        entropy = calculate_entropy(scores)
        
        ai_conf = ss.get("confidence", 0)
        ai_label = ss.get("label", "neutral")
        
        qualified_data.append({
            "raw_text_id": row["raw_text_id"],
            "entity_name": entity_name,
            "pseudo_label": ai_label,
            "ground_truth_label": "",
            "ai_confidence": round(ai_conf, 3),
            "entropy": round(entropy, 3),
            "media": rt.get("resolved_domain") or "unknown",
            "quality_score": meta.get("quality_score", 0),
            "context_text": ctx_text.replace("\n", " ").strip(),
            "article_text": full_text.replace("\n", " ").strip(),
            "source_url": source_url,
            "content_hash": rt.get("content_hash"),
            "entity_id": row["entity_id"]
        })

    logger.info(f"Data yang lolos Quality Filter: {len(qualified_data)}")
    if not qualified_data: return

    # Balancing & Pair Deduplication
    media_counter = Counter()
    entity_counter = Counter()
    
    random.shuffle(qualified_data)
    
    final_dataset = []
    seen_pairs = set()
    
    for item in qualified_data:
        pair_key = (item["content_hash"], item["entity_id"])
        if pair_key in seen_pairs: 
            audit["duplicate_pair"] += 1
            continue
            
        media = item["media"]
        entity = item["entity_name"]
        
        if media_counter[media] >= MAX_PER_MEDIA: 
            audit["media_limit"] += 1
            continue
        if entity_counter[entity] >= MAX_PER_ENTITY: 
            audit["entity_limit"] += 1
            continue
        
        seen_pairs.add(pair_key)
        media_counter[media] += 1
        entity_counter[entity] += 1
        final_dataset.append(item)
        
    # ==========================================
    # 1. EXPORT KE EXCEL (.xlsx) UNTUK MANUSIA
    # ==========================================
    df = pd.DataFrame(final_dataset)
    
    try:
        with pd.ExcelWriter(output_xlsx, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Gold Dataset')
            
            workbook = writer.book
            worksheet = writer.sheets['Gold Dataset']
            
            # Header Bold
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = Alignment(horizontal="center", vertical="center")
            for col in worksheet.iter_cols(1, worksheet.max_column, 1, 1):
                for cell in col:
                    cell.font = header_font
                    cell.alignment = header_fill
                    
            # Lebar Kolom
            col_widths = {
                'A': 35, 'B': 25, 'C': 15, 'D': 20, 'E': 15, 'F': 15,
                'G': 20, 'H': 15, 'I': 60, 'J': 100, 'K': 40, 'L': 35, 'M': 40
            }
            for col_letter, width in col_widths.items():
                worksheet.column_dimensions[col_letter].width = width
                
            # Wrap Text
            wrap_alignment = Alignment(wrap_text=True, vertical='top')
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=9, max_col=10):
                for cell in row:
                    cell.alignment = wrap_alignment
                    
        logger.info(f"[1/2] Excel untuk Human Labeler tersimpan: {output_xlsx}")
    except Exception as e:
        logger.error(f"Gagal menyimpan file Excel: {e}")

    # ==========================================
    # 2. EXPORT KE JSONL (.jsonl) UNTUK AI
    # ==========================================
    try:
        with open(output_jsonl, 'w', encoding='utf-8') as f:
            for item in final_dataset:
                # Siapkan format JSON yang hemat token
                ai_item = {
                    "raw_text_id": item["raw_text_id"],
                    "entity_name": item["entity_name"],
                    "pseudo_label": item["pseudo_label"],
                    "ground_truth_label": "", # Kosong, akan diisi model nanti
                    "context_text": item["context_text"],
                    # Potong article_text agar tidak boros token saat training
                    "article_text": item["article_text"][:MAX_ARTICLE_CHARS_AI],
                    "source_url": item["source_url"]
                }
                f.write(json.dumps(ai_item, ensure_ascii=False) + '\n')
        logger.info(f"[2/2] JSONL untuk AI Fine-Tuning tersimpan: {output_jsonl}")
    except Exception as e:
        logger.error(f"Gagal menyimpan file JSONL: {e}")
        
    logger.info("=" * 60)
    logger.info("GOLD STANDARD CURATION SELESAI!")
    logger.info(f"Total Data Diekspor : {len(final_dataset)} baris")
    logger.info("-" * 60)
    
    # PRINT AUDIT REPORT
    logger.info("=========== FILTER AUDIT REPORT ===========")
    logger.info(f"Raw Contexts               : {audit['raw_contexts']}")
    logger.info(f"Missing sentiment          : {audit['missing_sentiment']}")
    logger.info(f"Article too short          : {audit['article_short']}")
    logger.info(f"Boilerplate fail           : {audit['boilerplate_fail']}")
    logger.info(f"Context too short          : {audit['context_short']}")
    logger.info(f"Duplicate (hash, entity)   : {audit['duplicate_pair']}")
    logger.info(f"Entity limit reached       : {audit['entity_limit']}")
    logger.info(f"Media limit reached        : {audit['media_limit']}")
    logger.info(f"Final Export               : {len(final_dataset)}")
    logger.info("=" * 60)
    
    logger.info("Distribusi Media (Top 5):")
    for m, c in media_counter.most_common(5): logger.info(f"  {m:20s}: {c}")
    logger.info("Distribusi Entity (Top 5):")
    for e, c in entity_counter.most_common(5): logger.info(f"  {e:20s}: {c}")
    logger.info("=" * 60)

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit", type=int, default=10000)
    args = parser.parse_args()
    main(limit=args.limit)